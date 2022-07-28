from types import NoneType

import openpyxl as xl
from io import BytesIO
import time
import webbrowser
from keyboard import press_and_release
import win32clipboard
from PIL import Image
import psutil


def send_to_clipboard(filepath):
    image = Image.open(filepath)

    output = BytesIO()
    image.convert("RGB").save(output, "BMP")
    data = output.getvalue()[14:]
    output.close()
    win32clipboard.OpenClipboard()
    win32clipboard.EmptyClipboard()
    win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
    win32clipboard.CloseClipboard()


def replace_reserved_chars_in_url(raw_message: str):
    raw_message = raw_message.replace('%', '%25')
    raw_message = raw_message.replace('&', '%26')
    raw_message = raw_message.replace('$', '%24')
    raw_message = raw_message.replace('=', '%3D')
    return raw_message


def check_if_whatsapp_is_running():
    for proc in psutil.process_iter():
        try:
            # Check if process name contains the given name string.
            if 'whatsapp' in proc.name().lower():
                return True
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass
    return False


if not check_if_whatsapp_is_running():
    print('WhatsApp is not running.')
    print('Opening installed WhatsApp')
    webbrowser.open('whatsapp://')
    time.sleep(15)
    print('WhatsApp opened')
else:
    print('WhatsApp is running.')

image_folder = 'imgs\\'

wb = xl.load_workbook('contacts.xlsx')
sheet = wb['Sheet1']

for row in range(2, sheet.max_row + 1):
    image_exists = False
    image_file_path = 'No image added'
    if type(sheet.cell(row, 1).value) is not NoneType:
        mobile_number = str(sheet.cell(row, 1).value) + str(sheet.cell(row, 2).value)
    else:
        mobile_number = sheet.cell(row, 2).value
    message = str(sheet.cell(row, 3).value)
    try:
        if type(sheet.cell(row, 4).value) is not NoneType:
            image_exists = True
            image_file_path = image_folder + sheet.cell(row, 4).value
            send_to_clipboard(image_file_path)
    except FileNotFoundError:
        image_file_path = 'No images found!'
    replace_reserved_chars_in_url(message)
    print(f'sending message, To: {mobile_number}, message: {message}, image location: {image_file_path}')
    webbrowser.open(f'whatsapp://send?phone={mobile_number}&text={replace_reserved_chars_in_url(message)}')
    time.sleep(3)
    if image_exists:
        press_and_release('ctrl+v')
        time.sleep(1)
    press_and_release('enter')
